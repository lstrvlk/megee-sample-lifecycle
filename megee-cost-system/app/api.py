from datetime import datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from typing import Dict
from urllib.parse import quote
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.config import get_settings
from app.cost_engine import CostEngine, money, snapshot_fields
from app.database import get_db
from app.models import (
    SKU,
    Assembly,
    CostDataSubmission,
    CostSnapshot,
    CostVersion,
    Mold,
    OutsourcedPart,
    Part,
    ProductionRecord,
    Quotation,
    Routing,
    StandardComponent,
)
from app.schemas import (
    BomWorksheetItem,
    BomWorksheetRequest,
    CatalogImportRequest,
    CostComputeRequest,
    CostVersionCreateRequest,
    CostVersionDiffRequest,
    MoldUpdateRequest,
    ProductionInput,
    ProductionPreviewRequest,
    ProductionSubmissionBatch,
    ProductionSubmissionCreate,
    QuotationGenerateRequest,
    SubmissionReviewRequest,
)

router = APIRouter()
SNAPSHOT_FIELDS = (
    "material_cost",
    "process_cost",
    "mold_cost",
    "assembly_cost",
    "standard_component_cost",
    "outsourced_cost",
    "packaging_cost",
    "overhead_cost",
    "total_cost",
)
BOM_HEADERS = (
    "SKU编号",
    "产品名称",
    "包装成本",
    "制造费率",
    "组件编号",
    "组件名称",
    "组件类型",
    "零件编号",
    "零件名称",
    "成本类型",
    "用量",
    "未税单价",
    "供应商",
    "良率",
    "数据来源/备注",
)


@router.get("/health")
def health() -> dict:
    return {"status": "ok", "service": get_settings().app_name}


@router.get("/catalog/skus")
def list_skus(db: Session = Depends(get_db)) -> dict:
    skus = db.scalars(select(SKU).order_by(SKU.created_at.desc())).all()
    return {
        "items": [
            {
                "sku_id": sku.sku_id,
                "name": sku.name,
                "version": sku.version,
                "status": sku.status,
                "packaging_cost": sku.packaging_cost,
                "overhead_rate": sku.overhead_rate,
                "assembly_count": len(sku.assemblies),
            }
            for sku in skus
        ]
    }


@router.get("/catalog/skus/{sku_id}/operations")
def list_sku_operations(sku_id: str, db: Session = Depends(get_db)) -> dict:
    rows = db.execute(
        select(Part.part_id, Part.name, Routing.process_type, Routing.cycle_time)
        .join(Assembly, Part.assembly_id == Assembly.assembly_id)
        .join(Routing, Routing.part_id == Part.part_id)
        .where(Assembly.sku_id == sku_id)
        .order_by(Part.part_id, Routing.sequence)
    ).all()
    return {
        "items": [
            {
                "part_id": part_id,
                "part_name": part_name,
                "process_type": process_type,
                "standard_cycle_time": cycle_time,
            }
            for part_id, part_name, process_type, cycle_time in rows
        ]
    }


@router.get("/costing/skus/{sku_id}/bom")
def get_bom_worksheet(sku_id: str, db: Session = Depends(get_db)) -> dict:
    sku = CostEngine(db).load_sku(sku_id)
    request = _worksheet_from_sku(sku)
    return {**request.model_dump(mode="json"), "preview": _preview_bom(request)}


@router.post("/costing/bom/preview")
def preview_bom_worksheet(payload: BomWorksheetRequest) -> dict:
    return _preview_bom(payload)


@router.post("/costing/bom/apply")
def apply_bom_worksheet(payload: BomWorksheetRequest, db: Session = Depends(get_db)) -> dict:
    existing = db.get(SKU, payload.sku_id)
    sku = existing or SKU(sku_id=payload.sku_id, name=payload.sku_name)
    if existing is None:
        db.add(sku)
    else:
        sku.version += 1
    sku.name = payload.sku_name
    sku.packaging_cost = payload.packaging_cost
    sku.overhead_rate = payload.overhead_rate

    retained_part_ids = {row.part_id for row in payload.items}
    for row in payload.items:
        assembly = db.get(Assembly, row.assembly_id)
        if assembly is not None and assembly.sku_id != payload.sku_id:
            raise HTTPException(status_code=409, detail=f"组件编号 {row.assembly_id} 已被其他产品使用")
        if assembly is None:
            assembly = Assembly(assembly_id=row.assembly_id, sku_id=payload.sku_id)
            db.add(assembly)
        assembly.name = row.assembly_name
        assembly.type = row.assembly_type

        part = db.get(Part, row.part_id)
        if part is not None and part.assembly_id != row.assembly_id:
            raise HTTPException(status_code=409, detail=f"零件编号 {row.part_id} 已属于其他组件")
        if part is None:
            part = Part(part_id=row.part_id, assembly_id=row.assembly_id)
            db.add(part)
        part.name = row.part_name
        part.type = row.cost_type
        part.quantity = row.quantity

        if row.cost_type == "manufactured":
            part.material_cost = row.unit_price
            part.standard_component_id = None
        elif row.cost_type == "standard":
            component_id = part.standard_component_id or f"STD-{row.part_id}"
            component = db.get(StandardComponent, component_id)
            if component is None:
                component = StandardComponent(component_id=component_id)
                db.add(component)
            component.name = row.part_name
            component.unit_price = row.unit_price
            component.supplier = row.supplier or "待补充"
            part.standard_component_id = component_id
        else:
            quote = db.scalar(
                select(OutsourcedPart).where(
                    OutsourcedPart.part_id == row.part_id,
                    OutsourcedPart.is_active.is_(True),
                )
            )
            if quote is None:
                quote = OutsourcedPart(
                    outsourced_id=f"OUT-{row.part_id}",
                    part_id=row.part_id,
                    process_type="outsourced",
                )
                db.add(quote)
            quote.supplier = row.supplier or "待补充"
            quote.unit_price = row.unit_price
            quote.yield_assumption = row.yield_rate
            quote.lead_time = quote.lead_time or 0
            quote.is_active = True

    try:
        db.flush()
        stale_parts = db.scalars(
            select(Part)
            .join(Assembly, Part.assembly_id == Assembly.assembly_id)
            .where(Assembly.sku_id == payload.sku_id, Part.part_id.not_in(retained_part_ids))
        ).all()
        for part in stale_parts:
            part.quantity = Decimal("0")
        db.flush()
        cost = CostEngine(db).compute_sku(payload.sku_id, "standard")
        latest = db.scalar(
            select(CostVersion)
            .where(CostVersion.sku_id == payload.sku_id)
            .order_by(CostVersion.created_at.desc())
        )
        version = CostVersion(
            version_id=f"CV-BOM-{uuid4().hex[:10].upper()}",
            sku_id=payload.sku_id,
            parent_version_id=latest.version_id if latest else None,
            version_type="adjusted",
            reason=f"{payload.reason} · {payload.updated_by}",
        )
        version.snapshot = CostSnapshot(**snapshot_fields(cost))
        db.add(version)
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="BOM保存失败，请检查产品、组件和零件编号") from exc
    return {
        "status": "applied",
        "sku_id": payload.sku_id,
        "sku_version": sku.version,
        "version_id": version.version_id,
        "bom_preview": _preview_bom(payload),
        "standard_cost": cost,
    }


@router.get("/costing/bom/template/{sku_id}")
def download_bom_template(sku_id: str, db: Session = Depends(get_db)) -> StreamingResponse:
    sku = CostEngine(db).load_sku(sku_id)
    return _workbook_response(_build_bom_workbook(_worksheet_from_sku(sku), report=False), f"{sku_id}-BOM导入模板.xlsx")


@router.get("/costing/bom/template")
def download_blank_bom_template() -> StreamingResponse:
    payload = BomWorksheetRequest(
        sku_id="SKU-请填写",
        sku_name="产品名称",
        items=[
            BomWorksheetItem(
                assembly_id="ASM-001",
                assembly_name="瓶体总成",
                assembly_type="bottle",
                part_id="PART-001",
                part_name="零件名称",
                cost_type="manufactured",
                quantity=Decimal("1"),
                unit_price=Decimal("0"),
                source_note="请填写报价单或核价日期",
            )
        ],
    )
    return _workbook_response(_build_bom_workbook(payload, report=False), "MEGEE-BOM导入模板.xlsx")


@router.get("/costing/bom/report/{sku_id}")
def download_bom_report(sku_id: str, db: Session = Depends(get_db)) -> StreamingResponse:
    sku = CostEngine(db).load_sku(sku_id)
    return _workbook_response(_build_bom_workbook(_worksheet_from_sku(sku), report=True), f"{sku_id}-BOM成本报告.xlsx")


@router.post("/costing/bom/import")
async def import_bom_workbook(file: UploadFile = File(...)) -> dict:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="请上传 .xlsx 格式的 BOM 文件")
    try:
        workbook = load_workbook(BytesIO(await file.read()), data_only=True)
        worksheet = workbook["BOM"] if "BOM" in workbook.sheetnames else workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("工作表为空")
        header = [str(value).strip() if value is not None else "" for value in rows[0]]
        positions = {name: header.index(name) for name in BOM_HEADERS if name in header}
        missing = [name for name in BOM_HEADERS if name not in positions]
        if missing:
            raise ValueError(f"缺少列：{'、'.join(missing)}")
        records = [row for row in rows[1:] if row[positions["零件编号"]] not in (None, "")]
        if not records:
            raise ValueError("没有可导入的 BOM 明细")
        first = records[0]
        payload = BomWorksheetRequest(
            sku_id=str(first[positions["SKU编号"]]).strip(),
            sku_name=str(first[positions["产品名称"]]).strip(),
            packaging_cost=_cell_decimal(first[positions["包装成本"]]),
            overhead_rate=_cell_decimal(first[positions["制造费率"]]),
            items=[
                BomWorksheetItem(
                    assembly_id=str(row[positions["组件编号"]]).strip(),
                    assembly_name=str(row[positions["组件名称"]]).strip(),
                    assembly_type=str(row[positions["组件类型"]] or "custom").strip(),
                    part_id=str(row[positions["零件编号"]]).strip(),
                    part_name=str(row[positions["零件名称"]]).strip(),
                    cost_type=str(row[positions["成本类型"]]).strip().lower(),
                    quantity=_cell_decimal(row[positions["用量"]]),
                    unit_price=_cell_decimal(row[positions["未税单价"]]),
                    supplier=str(row[positions["供应商"]] or "").strip(),
                    yield_rate=_cell_decimal(row[positions["良率"]], Decimal("1")),
                    source_note=str(row[positions["数据来源/备注"]] or "").strip(),
                )
                for row in records
            ],
        )
    except (ValueError, TypeError, KeyError) as exc:
        raise HTTPException(status_code=422, detail=f"Excel 解析失败：{exc}") from exc
    return {**payload.model_dump(mode="json"), "preview": _preview_bom(payload)}


@router.get("/cost/versions/{sku_id}")
def list_cost_versions(sku_id: str, db: Session = Depends(get_db)) -> dict:
    if db.get(SKU, sku_id) is None:
        raise HTTPException(status_code=404, detail=f"SKU {sku_id} not found")
    versions = db.scalars(
        select(CostVersion)
        .where(CostVersion.sku_id == sku_id)
        .order_by(CostVersion.created_at.desc())
    ).all()
    return {
        "items": [
            {
                "version_id": version.version_id,
                "parent_version_id": version.parent_version_id,
                "version_type": version.version_type,
                "reason": version.reason,
                "created_at": version.created_at,
                "total_cost": version.snapshot.total_cost if version.snapshot else None,
            }
            for version in versions
        ]
    }


@router.get("/quotations/{sku_id}")
def list_quotations(sku_id: str, db: Session = Depends(get_db)) -> dict:
    quotations = db.scalars(
        select(Quotation)
        .where(Quotation.sku_id == sku_id)
        .order_by(Quotation.created_at.desc())
        .limit(20)
    ).all()
    return {
        "items": [
            {
                "quotation_no": quote.quotation_no,
                "customer_name": quote.customer_name,
                "quantity": quote.quantity,
                "currency": quote.currency,
                "unit_cost": quote.unit_cost,
                "unit_price": quote.unit_price,
                "total_price": quote.total_price,
                "created_at": quote.created_at,
            }
            for quote in quotations
        ]
    }


@router.post("/catalog/import", status_code=status.HTTP_200_OK)
def import_catalog(payload: CatalogImportRequest, db: Session = Depends(get_db)) -> dict:
    groups = (
        (payload.skus, SKU),
        (payload.assemblies, Assembly),
        (payload.standard_components, StandardComponent),
        (payload.parts, Part),
        (payload.routings, Routing),
        (payload.molds, Mold),
        (payload.outsourced_parts, OutsourcedPart),
    )
    counts: Dict[str, int] = {}
    try:
        for items, model in groups:
            counts[model.__tablename__] = len(items)
            for item in items:
                db.merge(model(**item.model_dump()))
            db.flush()
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="Catalog import violates a data relationship") from exc
    return {"status": "imported", "counts": counts}


@router.post("/sku/calculate")
def calculate_sku(payload: CostComputeRequest, db: Session = Depends(get_db)) -> dict:
    return CostEngine(db).compute_sku(payload.sku_id, payload.stage, payload.include_optional)


@router.post("/cost/compute")
def compute_cost(payload: CostComputeRequest, db: Session = Depends(get_db)) -> dict:
    return CostEngine(db).compute_sku(payload.sku_id, payload.stage, payload.include_optional)


@router.post("/production/input", status_code=status.HTTP_201_CREATED)
def input_production(payload: ProductionInput, db: Session = Depends(get_db)) -> dict:
    if db.get(ProductionRecord, payload.record_id) is not None:
        raise HTTPException(status_code=409, detail=f"Production record {payload.record_id} already exists")
    if db.get(Part, payload.part_id) is None:
        raise HTTPException(status_code=404, detail=f"Part {payload.part_id} not found")
    record = ProductionRecord(**payload.model_dump())
    db.add(record)
    db.commit()
    db.refresh(record)
    return {
        "record_id": record.record_id,
        "part_id": record.part_id,
        "source": record.source,
        "yield_rate": money(Decimal(record.good_qty) / Decimal(record.qty)),
        "recorded_at": record.recorded_at,
    }


@router.post("/cost-data/submissions", status_code=status.HTTP_201_CREATED)
def create_cost_data_submission(
    payload: ProductionSubmissionCreate, db: Session = Depends(get_db)
) -> dict:
    _validate_production_for_sku(db, payload.sku_id, payload.production)
    submission = _new_production_submission(
        payload.sku_id,
        payload.submitted_by,
        payload.source_mode,
        payload.production,
    )
    db.add(submission)
    db.commit()
    db.refresh(submission)
    return _submission_response(submission)


@router.post("/cost-data/import", status_code=status.HTTP_201_CREATED)
def import_cost_data(payload: ProductionSubmissionBatch, db: Session = Depends(get_db)) -> dict:
    record_ids = [row.record_id for row in payload.rows]
    if len(record_ids) != len(set(record_ids)):
        raise HTTPException(status_code=422, detail="Imported rows contain duplicate record_id values")
    for row in payload.rows:
        _validate_production_for_sku(db, payload.sku_id, row)
    submissions = [
        _new_production_submission(payload.sku_id, payload.submitted_by, "import", row)
        for row in payload.rows
    ]
    db.add_all(submissions)
    db.commit()
    return {
        "status": "pending_approval",
        "count": len(submissions),
        "submission_ids": [item.submission_id for item in submissions],
    }


@router.get("/cost-data/submissions")
def list_cost_data_submissions(
    sku_id: str, submission_status: str = "pending", db: Session = Depends(get_db)
) -> dict:
    statement = select(CostDataSubmission).where(CostDataSubmission.sku_id == sku_id)
    if submission_status != "all":
        statement = statement.where(CostDataSubmission.status == submission_status)
    items = db.scalars(statement.order_by(CostDataSubmission.submitted_at.desc()).limit(100)).all()
    return {"items": [_submission_response(item) for item in items]}


@router.post("/cost-data/preview")
def preview_cost_data(payload: ProductionPreviewRequest, db: Session = Depends(get_db)) -> dict:
    _validate_production_for_sku(
        db, payload.sku_id, payload.production, check_pending=False, check_record=False
    )
    part = db.get(Part, payload.production.part_id)
    route = db.scalar(
        select(Routing).where(
            Routing.part_id == payload.production.part_id,
            Routing.process_type == payload.production.process_type,
        )
    )
    current_cost = CostEngine(db).compute_sku(payload.sku_id, payload.production.source)
    projected_cost = CostEngine(db, payload.production).compute_sku(
        payload.sku_id, payload.production.source
    )

    existing = db.scalars(
        select(ProductionRecord).where(
            ProductionRecord.part_id == payload.production.part_id,
            ProductionRecord.process_type == payload.production.process_type,
            ProductionRecord.source == payload.production.source,
        )
    ).all()
    existing_qty = sum((item.qty for item in existing), 0)
    existing_good = sum((item.good_qty for item in existing), 0)
    existing_cycle_total = sum(
        (Decimal(item.cycle_time_actual) * item.qty for item in existing), Decimal("0")
    )
    projected_qty = existing_qty + payload.production.qty
    projected_good = existing_good + payload.production.good_qty
    projected_cycle = (
        existing_cycle_total
        + Decimal(payload.production.cycle_time_actual) * payload.production.qty
    ) / projected_qty
    projected_yield = Decimal(projected_good) / Decimal(projected_qty)
    hourly_rate = Decimal(route.machine_rate_per_hour) + (
        Decimal(route.labor_rate_per_hour) * route.labor_count
    )
    raw_process_cost = hourly_rate * Decimal(route.cycle_time) / Decimal("3600")
    process_factor = projected_cycle / Decimal(route.cycle_time)
    yield_cost_factor = Decimal(projected_qty) / Decimal(projected_good)
    projected_process_cost = raw_process_cost * process_factor * yield_cost_factor
    mold_unit_cost = sum(
        (CostEngine._mold_unit_cost(mold) for mold in part.molds), Decimal("0")
    )

    return {
        "operation": {
            "part_id": part.part_id,
            "part_name": part.name,
            "process_type": route.process_type,
            "material_cost": money(Decimal(part.material_cost)),
            "standard_cycle_time": route.cycle_time,
            "machine": route.machine,
            "machine_rate_per_hour": route.machine_rate_per_hour,
            "labor_count": route.labor_count,
            "labor_rate_per_hour": route.labor_rate_per_hour,
            "standard_yield": route.yield_rate,
            "mold_unit_cost": money(mold_unit_cost),
        },
        "aggregation": {
            "existing_record_count": len(existing),
            "existing_qty": existing_qty,
            "candidate_qty": payload.production.qty,
            "projected_qty": projected_qty,
            "projected_good_qty": projected_good,
            "projected_cycle_time": money(projected_cycle),
            "projected_yield": money(projected_yield),
        },
        "formula": {
            "hourly_rate": money(hourly_rate),
            "raw_process_cost": money(raw_process_cost),
            "process_factor": money(process_factor),
            "yield_cost_factor": money(yield_cost_factor),
            "projected_process_cost": money(projected_process_cost),
        },
        "sku_impact": {
            "stage": payload.production.source,
            "current_total_cost": current_cost["total_cost"],
            "projected_total_cost": projected_cost["total_cost"],
            "delta": money(
                Decimal(projected_cost["total_cost"]) - Decimal(current_cost["total_cost"])
            ),
        },
    }


@router.post("/cost-data/submissions/{submission_id}/approve")
def approve_cost_data_submission(
    submission_id: str, payload: SubmissionReviewRequest, db: Session = Depends(get_db)
) -> dict:
    submission = _pending_submission(db, submission_id)
    production = ProductionInput.model_validate(submission.payload)
    _validate_production_for_sku(db, submission.sku_id, production, check_pending=False)
    if db.get(ProductionRecord, production.record_id) is not None:
        raise HTTPException(status_code=409, detail=f"Production record {production.record_id} already exists")

    before = CostEngine(db).compute_sku(submission.sku_id, production.source)
    record = ProductionRecord(**production.model_dump())
    db.add(record)
    db.flush()
    after = CostEngine(db).compute_sku(submission.sku_id, production.source)

    parent = db.scalar(
        select(CostVersion)
        .where(CostVersion.sku_id == submission.sku_id)
        .order_by(CostVersion.created_at.desc())
        .limit(1)
    )
    version_id = f"CV-APP-{uuid4().hex[:10].upper()}"
    version = CostVersion(
        version_id=version_id,
        sku_id=submission.sku_id,
        parent_version_id=parent.version_id if parent else None,
        version_type=production.source,
        reason=f"Approved field data {submission.submission_id}",
    )
    version.snapshot = CostSnapshot(**snapshot_fields(after))
    db.add(version)
    db.flush()

    submission.status = "approved"
    submission.reviewed_by = payload.reviewed_by
    submission.reviewed_at = datetime.now(timezone.utc)
    submission.review_comment = payload.comment
    submission.applied_record_id = production.record_id
    submission.applied_version_id = version_id
    db.commit()
    return {
        "submission": _submission_response(submission),
        "cost_impact": {
            "stage": production.source,
            "before": before["total_cost"],
            "after": after["total_cost"],
            "delta": money(Decimal(after["total_cost"]) - Decimal(before["total_cost"])),
        },
    }


@router.post("/cost-data/submissions/{submission_id}/reject")
def reject_cost_data_submission(
    submission_id: str, payload: SubmissionReviewRequest, db: Session = Depends(get_db)
) -> dict:
    submission = _pending_submission(db, submission_id)
    submission.status = "rejected"
    submission.reviewed_by = payload.reviewed_by
    submission.reviewed_at = datetime.now(timezone.utc)
    submission.review_comment = payload.comment
    db.commit()
    return _submission_response(submission)


@router.post("/mold/update")
def update_mold(payload: MoldUpdateRequest, db: Session = Depends(get_db)) -> dict:
    mold = db.get(Mold, payload.mold_id)
    if mold is None:
        raise HTTPException(status_code=404, detail=f"Mold {payload.mold_id} not found")
    if payload.actual_output is not None:
        mold.actual_output = payload.actual_output
    elif payload.output_increment is not None:
        mold.actual_output += payload.output_increment
    if payload.lifecycle_status is not None:
        mold.lifecycle_status = payload.lifecycle_status
    db.commit()
    db.refresh(mold)

    remaining_output = max(mold.planned_output - mold.actual_output, 0)
    remaining_value = (
        Decimal(mold.cost) * Decimal(remaining_output) / Decimal(mold.planned_output)
        if mold.planned_output > 0 and mold.lifecycle_status != "end"
        else Decimal("0")
    )
    return {
        "mold_id": mold.mold_id,
        "actual_output": mold.actual_output,
        "planned_output": mold.planned_output,
        "remaining_output": remaining_output,
        "remaining_value": money(remaining_value),
        "lifecycle_status": mold.lifecycle_status,
    }


@router.post("/cost/version/create", status_code=status.HTTP_201_CREATED)
def create_cost_version(payload: CostVersionCreateRequest, db: Session = Depends(get_db)) -> dict:
    if db.get(CostVersion, payload.version_id) is not None:
        raise HTTPException(status_code=409, detail=f"Cost version {payload.version_id} already exists")
    if payload.parent_version_id:
        parent = db.get(CostVersion, payload.parent_version_id)
        if parent is None:
            raise HTTPException(status_code=404, detail="Parent cost version not found")
        if parent.sku_id != payload.sku_id:
            raise HTTPException(status_code=422, detail="Parent cost version belongs to another SKU")

    cost = CostEngine(db).compute_sku(payload.sku_id, payload.stage, payload.include_optional)
    version = CostVersion(
        version_id=payload.version_id,
        sku_id=payload.sku_id,
        parent_version_id=payload.parent_version_id,
        version_type=payload.version_type,
        reason=payload.reason,
    )
    version.snapshot = CostSnapshot(**snapshot_fields(cost))
    db.add(version)
    db.commit()
    db.refresh(version)
    return {
        "version_id": version.version_id,
        "sku_id": version.sku_id,
        "parent_version_id": version.parent_version_id,
        "version_type": version.version_type,
        "created_at": version.created_at,
        "snapshot": snapshot_fields(cost),
    }


@router.post("/cost/version/diff")
def diff_cost_versions(payload: CostVersionDiffRequest, db: Session = Depends(get_db)) -> dict:
    old = _get_version_with_snapshot(db, payload.from_version_id)
    new = _get_version_with_snapshot(db, payload.to_version_id)
    if old.sku_id != new.sku_id:
        raise HTTPException(status_code=422, detail="Cost versions belong to different SKUs")

    differences = {}
    for field_name in SNAPSHOT_FIELDS:
        old_value = Decimal(getattr(old.snapshot, field_name))
        new_value = Decimal(getattr(new.snapshot, field_name))
        delta = new_value - old_value
        percent = (delta / old_value * Decimal("100")) if old_value else None
        differences[field_name] = {
            "from": money(old_value),
            "to": money(new_value),
            "delta": money(delta),
            "change_percent": money(percent) if percent is not None else None,
        }
    return {
        "sku_id": old.sku_id,
        "from_version_id": old.version_id,
        "to_version_id": new.version_id,
        "differences": differences,
    }


@router.post("/quotation/generate", status_code=status.HTTP_201_CREATED)
def generate_quotation(payload: QuotationGenerateRequest, db: Session = Depends(get_db)) -> dict:
    cost = CostEngine(db).compute_sku(payload.sku_id, payload.stage, payload.include_optional)
    unit_cost = Decimal(cost["total_cost"])
    risk_items = []

    if payload.is_new_product:
        risk_items.append(("new_product", Decimal("0.05")))
    if payload.is_new_customer:
        risk_items.append(("new_customer", Decimal("0.05")))
    if payload.is_small_batch:
        risk_items.append(("small_batch", Decimal("0.10")))

    outsourced_cost = Decimal(cost["outsourced_cost"])
    outsourced_ratio = outsourced_cost / unit_cost if unit_cost else Decimal("0")
    outsourced_risk = payload.outsourced_risk_rate
    if outsourced_risk is None:
        outsourced_risk = _outsourced_risk_rate(outsourced_ratio)
    if outsourced_cost > 0 and outsourced_risk > 0:
        risk_items.append(("outsourced", Decimal(outsourced_risk)))

    risk_rate = sum((rate for _, rate in risk_items), Decimal("0"))
    risk_adjusted_cost = unit_cost * (Decimal("1") + risk_rate)
    unit_price = risk_adjusted_cost / (Decimal("1") - payload.target_margin)
    total_price = unit_price * payload.quantity
    china_standard_time = timezone(timedelta(hours=8))
    now = datetime.now(china_standard_time)
    quotation_id = uuid4().hex
    quotation_no = f"QT-{now:%Y%m%d}-{quotation_id[:8].upper()}"
    quotation = Quotation(
        quotation_id=quotation_id,
        quotation_no=quotation_no,
        sku_id=payload.sku_id,
        customer_name=payload.customer_name,
        quantity=payload.quantity,
        currency=payload.currency,
        unit_cost=money(unit_cost),
        risk_rate=risk_rate,
        target_margin=payload.target_margin,
        unit_price=money(unit_price),
        total_price=money(total_price),
    )
    db.add(quotation)
    db.commit()
    return {
        "quotation_id": quotation.quotation_id,
        "quotation_no": quotation.quotation_no,
        "sku_id": quotation.sku_id,
        "customer_name": quotation.customer_name,
        "quantity": quotation.quantity,
        "currency": quotation.currency,
        "unit_cost": quotation.unit_cost,
        "risk_items": [{"type": name, "rate": rate} for name, rate in risk_items],
        "risk_rate": money(risk_rate),
        "target_margin": money(Decimal(payload.target_margin)),
        "unit_price": quotation.unit_price,
        "total_price": quotation.total_price,
        "created_at": quotation.created_at,
    }


def _get_version_with_snapshot(db: Session, version_id: str) -> CostVersion:
    version = db.scalar(select(CostVersion).where(CostVersion.version_id == version_id))
    if version is None or version.snapshot is None:
        raise HTTPException(status_code=404, detail=f"Cost version {version_id} not found")
    return version


def _outsourced_risk_rate(ratio: Decimal) -> Decimal:
    if ratio >= Decimal("0.50"):
        return Decimal("0.08")
    if ratio >= Decimal("0.25"):
        return Decimal("0.05")
    if ratio > 0:
        return Decimal("0.03")
    return Decimal("0")


def _validate_production_for_sku(
    db: Session,
    sku_id: str,
    production: ProductionInput,
    check_pending: bool = True,
    check_record: bool = True,
) -> None:
    operation = db.execute(
        select(Part.part_id)
        .join(Assembly, Part.assembly_id == Assembly.assembly_id)
        .join(Routing, Routing.part_id == Part.part_id)
        .where(
            Assembly.sku_id == sku_id,
            Part.part_id == production.part_id,
            Routing.process_type == production.process_type,
        )
    ).first()
    if operation is None:
        raise HTTPException(
            status_code=422,
            detail=f"Operation {production.part_id}/{production.process_type} does not belong to SKU {sku_id}",
        )
    if check_record and db.get(ProductionRecord, production.record_id) is not None:
        raise HTTPException(status_code=409, detail=f"Production record {production.record_id} already exists")
    if check_pending:
        pending = db.scalars(
            select(CostDataSubmission).where(
                CostDataSubmission.sku_id == sku_id,
                CostDataSubmission.status == "pending",
            )
        ).all()
        if any(item.payload.get("record_id") == production.record_id for item in pending):
            raise HTTPException(
                status_code=409,
                detail=f"Production record {production.record_id} is already pending approval",
            )


def _new_production_submission(
    sku_id: str,
    submitted_by: str,
    source_mode: str,
    production: ProductionInput,
) -> CostDataSubmission:
    return CostDataSubmission(
        submission_id=f"CDS-{uuid4().hex[:12].upper()}",
        sku_id=sku_id,
        data_type="production",
        source_mode=source_mode,
        submitted_by=submitted_by,
        payload=production.model_dump(mode="json"),
        status="pending",
    )


def _pending_submission(db: Session, submission_id: str) -> CostDataSubmission:
    submission = db.get(CostDataSubmission, submission_id)
    if submission is None:
        raise HTTPException(status_code=404, detail=f"Submission {submission_id} not found")
    if submission.status != "pending":
        raise HTTPException(
            status_code=409,
            detail=f"Submission {submission_id} has already been {submission.status}",
        )
    return submission


def _submission_response(submission: CostDataSubmission) -> dict:
    return {
        "submission_id": submission.submission_id,
        "sku_id": submission.sku_id,
        "data_type": submission.data_type,
        "source_mode": submission.source_mode,
        "submitted_by": submission.submitted_by,
        "payload": submission.payload,
        "status": submission.status,
        "submitted_at": submission.submitted_at,
        "reviewed_by": submission.reviewed_by,
        "reviewed_at": submission.reviewed_at,
        "review_comment": submission.review_comment,
        "applied_record_id": submission.applied_record_id,
        "applied_version_id": submission.applied_version_id,
    }


def _worksheet_from_sku(sku: SKU) -> BomWorksheetRequest:
    items = []
    for assembly in sorted(sku.assemblies, key=lambda item: item.assembly_id):
        for part in sorted(assembly.parts, key=lambda item: item.part_id):
            if Decimal(part.quantity) <= 0:
                continue
            unit_price = Decimal(part.material_cost)
            supplier = ""
            yield_rate = Decimal("1")
            source_note = "系统材料成本"
            if part.type == "standard" and part.standard_component is not None:
                unit_price = Decimal(part.standard_component.unit_price)
                supplier = part.standard_component.supplier
                source_note = "标准件采购价"
            elif part.type == "outsourced":
                active = [quote for quote in part.outsourced_quotes if quote.is_active]
                quote = min(
                    active,
                    key=lambda item: Decimal(item.unit_price) / Decimal(item.yield_assumption),
                    default=None,
                )
                if quote is not None:
                    unit_price = Decimal(quote.unit_price)
                    supplier = quote.supplier
                    yield_rate = Decimal(quote.yield_assumption)
                    source_note = "外协有效报价"
            items.append(
                BomWorksheetItem(
                    assembly_id=assembly.assembly_id,
                    assembly_name=assembly.name,
                    assembly_type=assembly.type,
                    part_id=part.part_id,
                    part_name=part.name,
                    cost_type=part.type,
                    quantity=Decimal(part.quantity),
                    unit_price=unit_price,
                    supplier=supplier,
                    yield_rate=yield_rate,
                    source_note=source_note,
                )
            )
    return BomWorksheetRequest(
        sku_id=sku.sku_id,
        sku_name=sku.name,
        packaging_cost=Decimal(sku.packaging_cost),
        overhead_rate=Decimal(sku.overhead_rate),
        items=items,
    )


def _preview_bom(payload: BomWorksheetRequest) -> dict:
    subtotals = {
        "manufactured": Decimal("0"),
        "standard": Decimal("0"),
        "outsourced": Decimal("0"),
    }
    lines = []
    for row in payload.items:
        adjusted_unit_price = row.unit_price
        formula = "用量 × 未税单价"
        if row.cost_type == "outsourced":
            adjusted_unit_price = row.unit_price / row.yield_rate
            formula = "用量 × 未税单价 ÷ 外协良率"
        line_cost = row.quantity * adjusted_unit_price
        subtotals[row.cost_type] += line_cost
        lines.append(
            {
                **row.model_dump(mode="json"),
                "adjusted_unit_price": money(adjusted_unit_price),
                "line_cost": money(line_cost),
                "formula": formula,
            }
        )
    direct_cost = sum(subtotals.values(), Decimal("0"))
    bom_total = direct_cost + payload.packaging_cost
    return {
        "lines": lines,
        "manufactured_cost": money(subtotals["manufactured"]),
        "standard_component_cost": money(subtotals["standard"]),
        "outsourced_cost": money(subtotals["outsourced"]),
        "packaging_cost": money(payload.packaging_cost),
        "direct_part_cost": money(direct_cost),
        "bom_total": money(bom_total),
        "overhead_rate": money(payload.overhead_rate),
    }


def _build_bom_workbook(payload: BomWorksheetRequest, report: bool) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    headers = [*BOM_HEADERS, "成本贡献"]
    sheet.append(headers)
    for index, row in enumerate(payload.items, start=2):
        sheet.append(
            [
                payload.sku_id,
                payload.sku_name,
                float(payload.packaging_cost),
                float(payload.overhead_rate),
                row.assembly_id,
                row.assembly_name,
                row.assembly_type,
                row.part_id,
                row.part_name,
                row.cost_type,
                float(row.quantity),
                float(row.unit_price),
                row.supplier,
                float(row.yield_rate),
                row.source_note,
                None,
            ]
        )
        sheet.cell(index, 16).value = f'=IF(J{index}="outsourced",K{index}*L{index}/N{index},K{index}*L{index})'

    dark_fill = PatternFill("solid", fgColor="1F2937")
    input_fill = PatternFill("solid", fgColor="EAF2FF")
    for cell in sheet[1]:
        cell.fill = dark_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for column in (3, 4, 11, 12, 13, 14, 15):
            row[column - 1].fill = input_fill
    for column in (3, 11, 12, 16):
        for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2, max_row=sheet.max_row):
            cell[0].number_format = '¥0.000000' if column in (3, 12, 16) else '0.000000'
    for cell in sheet.iter_cols(min_col=4, max_col=4, min_row=2, max_row=sheet.max_row):
        cell[0].number_format = "0.0%"
    for cell in sheet.iter_cols(min_col=14, max_col=14, min_row=2, max_row=sheet.max_row):
        cell[0].number_format = "0.0%"
    widths = (16, 24, 12, 12, 17, 20, 12, 19, 25, 14, 10, 13, 20, 10, 26, 14)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:P{sheet.max_row}"

    notes = workbook.create_sheet("填写说明")
    notes.append(["MEGEE BOM 成本导入说明"])
    notes.append(["蓝色单元格为需要核对或填写的真实参数。成本类型只能填写 manufactured、standard、outsourced。"])
    notes.append(["自制件成本贡献 = 用量 × 材料未税单价；标准件 = 用量 × 采购未税单价；外协 = 用量 × 未税单价 ÷ 良率。"])
    notes.append(["制造费率填写小数，例如 8.5% 填写为 0.085。导入后先试算，点击确认采用后才写入系统。"])
    notes.column_dimensions["A"].width = 110
    notes["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    notes["A1"].fill = dark_fill
    for row in range(2, 5):
        notes[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    if report:
        preview = _preview_bom(payload)
        summary = workbook.create_sheet("成本汇总", 0)
        summary.append(["MEGEE BOM 成本报告", payload.sku_id, payload.sku_name])
        summary.append(["成本项目", "单位成本（CNY/件）", "计算口径"])
        summary.append(["自制材料", float(Decimal(preview["manufactured_cost"])), "自制件用量 × 材料未税单价"])
        summary.append(["标准件", float(Decimal(preview["standard_component_cost"])), "标准件用量 × 采购未税单价"])
        summary.append(["外协件", float(Decimal(preview["outsourced_cost"])), "用量 × 外协未税单价 ÷ 良率"])
        summary.append(["包装", float(payload.packaging_cost), "产品包装成本"])
        summary.append(["BOM成本合计", float(Decimal(preview["bom_total"])), "以上项目合计，不含工艺、模具和制造费"])
        summary.column_dimensions["A"].width = 22
        summary.column_dimensions["B"].width = 22
        summary.column_dimensions["C"].width = 48
        summary.merge_cells("A1:C1")
        summary["A1"].fill = dark_fill
        summary["A1"].font = Font(color="FFFFFF", bold=True, size=16)
        summary["A1"].alignment = Alignment(horizontal="center")
        for cell in summary[2]:
            cell.fill = PatternFill("solid", fgColor="D9EAD3")
            cell.font = Font(bold=True)
        for row in range(3, 8):
            summary.cell(row, 2).number_format = '¥0.000000'
        summary["A7"].font = Font(bold=True)
        summary["B7"].font = Font(bold=True)
    return workbook


def _workbook_response(workbook: Workbook, filename: str) -> StreamingResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


def _cell_decimal(value: object, default: Decimal = Decimal("0")) -> Decimal:
    if value in (None, ""):
        return default
    return Decimal(str(value))
